# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import openpyxl
from Levenshtein import distance as dst
from Levenshtein import ratio as rt

class GetDataFromExcel():

    @staticmethod
    def rmv(string):
        """Python3 code to remove whitespace"""
        return "".join(str(string).split()).lower().replace(".", "")

    @staticmethod
    def matchlev(str1, str2, r=0.9, d=5):
        """ Метод для сравнения процента совпадения двух строк"""

        if not str1 or not str2:
            return False

        current_ratio = rt(str1.lower(), str2.lower())
        current_dist = dst(str1.lower(), str2.lower())
        if current_ratio >= r and current_dist <= d:
            return True

    @staticmethod
    def ifnull(var, val):
        """ Python equivalent for MySQL's IFNULL """

        if var is None or var.replace(" ", "") == "":
            return val
        return var


    def open_xlsx(self, file):

        self.path_xlsx = file
        self.wb_obj = openpyxl.load_workbook(self.path_xlsx)
        self.sheet_obj = self.wb_obj.active
        self.mCol = self.sheet_obj.max_column  # номер последней колонки
        self.mRow = self.sheet_obj.max_row  # номер последней строки

    def find_headers(self, headers, param=0):
        """ определение номера строки шапки таблицы"""

        sheet = self.sheet_obj
        for i in range(1, self.mRow + 1):
            row_str = " ".join([str(sheet.cell(row=i, column=a).value).lower() for a in range(1, self.mCol + 1)
                               if sheet.cell(row=i, column=a).value != None])
            if len(headers) == len([a for a in headers if a.lower() in row_str]):
                if param == 0:
                    res = {sheet.cell(row=i, column=a).value: (i,a) for a in range(1, self.mCol + 1)
                            if len([w for w in headers if self.matchlev(sheet.cell(row=i, column=a).value, w)]) > 0}
                else:
                    res = {sheet.cell(row=i, column=a).value: (i, a) for a in range(1, self.mCol + 1)
                           if len([w for w in headers if w in self.ifnull(sheet.cell(row=i, column=a).value, "")]) > 0}
                return res
        return {}

    def find_next_word(self, word, key=None):
        sheet = self.sheet_obj
        for i in range(1, self.mRow + 1):
            row_str = " ".join([sheet.cell(row=i, column=a).value.lower() for a in range(1, self.mCol + 1)
                               if sheet.cell(row=i, column=a).value != None])
            if word.lower() in row_str and self.ifnull(key, "") in row_str:
                s = [w for w in row_str.split(word)[1].split(" ") if self.ifnull(w, "") not in ["", ".", "-", ":"]][0]
                return s
        return "not found"


    def find_end_row(self, headers):
        for k, v in headers.items():
            start = v[0]+1
            for i in range(start, self.mRow):
                if len([key for key, val in headers.items() if self.sheet_obj.cell(row=i, column=val[1]).value]) < 2:
                    return start, i
            return start, self.mRow

    def gather_table_data(self, headers):
        startData, endData = self.find_end_row(headers)
        data = []
        for i in range(startData, endData+1):
            if int(self.sheet_obj.cell(row=i, column=1).fill.start_color.index) > 0:
                print("row ", i, "skipped, because has colour")
                continue
            data.append({key: str(self.sheet_obj.cell(row=i, column=val[1]).value) for key, val in headers.items()})
        return data

# cls = GetDataFromExcel(r"C:\Users\Yernur\Dropbox\PC\Downloads\download6X1v4172917\_____ ____ 1 __.xlsx")
# hd = cls.find_headers(["код", "наименование", "кол-во", "цена", "сумма"])
# print(cls.gather_table_data(hd))
#
# iik = cls.find_next_word("иик")
# kbe = cls.find_next_word("кбе")
# print(iik, kbe)
